import io
import openpyxl
from urllib.parse import quote
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Dict, Tuple
from fastapi import Depends, FastAPI, Form, Request, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.ext.asyncio import AsyncSession
from app import auth, crud, special_funcs
from app.db_engine import get_db, init_db


@asynccontextmanager
async def lifespan(_: FastAPI):
    await init_db()
    yield


app = FastAPI(title="Finance Flow API", version="1.0.0", lifespan=lifespan)
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
BASE_DIR = Path(__file__).resolve().parent.parent

COOKIE_ACCESS = "access_token"
COOKIE_REFRESH = "refresh_token"


def set_auth_cookies(response: Response, tokens: Dict[str, str]) -> None:
    response.set_cookie(COOKIE_ACCESS, tokens["access_token"], httponly=True, samesite="lax")
    response.set_cookie(COOKIE_REFRESH, tokens["refresh_token"], httponly=True, samesite="lax")


def delete_auth_cookies(response: Response) -> None:
    response.delete_cookie(COOKIE_ACCESS)
    response.delete_cookie(COOKIE_REFRESH)


async def get_user(request: Request, db: AsyncSession) -> Tuple:
    return await auth.get_authenticated_user(
        db,
        request.cookies.get(COOKIE_ACCESS),
        request.cookies.get(COOKIE_REFRESH),
    )


@app.get("/", response_class=HTMLResponse)
async def home(request: Request, db: AsyncSession = Depends(get_db)):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login")
    response = templates.TemplateResponse(request, "index.html", {"user": user})
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.get("/register")
async def register_form(request: Request):
    return templates.TemplateResponse(request, "register.html")


@app.post("/register")
async def register(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    username = username.strip()
    user = await crud.create_user(db, username, email, password)
    if user:
        return RedirectResponse("/login", status_code=302)
    return templates.TemplateResponse(request, "register.html", {"error": "Пользователь уже существует"})


@app.get("/login")
async def login_form(request: Request):
    return templates.TemplateResponse(request, "login.html")


@app.post("/login")
async def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    username = username.strip()
    user = await crud.authenticate_user(db, username, password)
    if not user:
        return templates.TemplateResponse(request, "login.html", {"error": "Неверные данные"})
    tokens = await auth.login_user(db, user)
    response = RedirectResponse("/", status_code=302)
    set_auth_cookies(response, tokens)
    return response


@app.get("/logout")
async def logout(request: Request, db: AsyncSession = Depends(get_db)):
    await auth.logout_user(db, request.cookies.get(COOKIE_ACCESS))
    response = RedirectResponse("/login")
    delete_auth_cookies(response)
    return response


@app.post("/import")
async def import_excel(
    request: Request,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return JSONResponse({"success": False, "error": "Не авторизован"}, status_code=401)
    if not file.filename.endswith((".xlsx", ".xls")):
        return JSONResponse({"success": False, "error": "Поддерживаются только файлы .xlsx / .xls"})
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        imported = 0
        skipped = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(cell is not None for cell in row):
                continue
            if len(row) < 3:
                skipped.append(f"строка {i}: мало столбцов")
                continue
            type_val = str(row[0]).strip().lower() if row[0] is not None else ""
            description = str(row[1]).strip() if row[1] is not None else ""
            amount_raw = row[2]
            currency = str(row[3]).strip().upper() if len(row) > 3 and row[3] else "RUB"
            if currency not in ("RUB", "USD", "EUR"):
                currency = "RUB"
            try:
                amount = float(amount_raw)
            except (ValueError, TypeError):
                skipped.append(f"строка {i}: некорректная сумма")
                continue
            if not description:
                skipped.append(f"строка {i}: пустое описание")
                continue
            if type_val in ("доход", "income", "доходы"):
                await crud.add_income(db, user.id, description, amount, currency)
                imported += 1
            elif type_val in ("расход", "expense", "расходы"):
                await crud.add_expense(db, user.id, description, amount, currency)
                imported += 1
            else:
                skipped.append(f"строка {i}: неизвестный тип «{row[0]}»")
        response_data: Dict = {"success": True, "imported": imported}
        if skipped:
            response_data["skipped"] = skipped
        resp = JSONResponse(response_data)
        if new_tokens:
            set_auth_cookies(resp, new_tokens)
        return resp
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=400)


@app.post("/add-income")
async def add_income(
    request: Request,
    description: str = Form(...),
    amount: float = Form(...),
    currency: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    await crud.add_income(db, user.id, description, amount, currency)
    response = RedirectResponse("/", status_code=302)
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.post("/add-expense")
async def add_expense(
    request: Request,
    description: str = Form(...),
    amount: float = Form(...),
    currency: str = Form(...),
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    await crud.add_expense(db, user.id, description, amount, currency)
    response = RedirectResponse("/", status_code=302)
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.get("/data")
async def show_data(
    request: Request,
    year: str = "",
    month_num: str = "",
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login")

    date = f"{year}-{month_num}" if year and month_num else ""
    export_period = f"{year}-{month_num}" if year and month_num else (year if year else "")
    incomes, expenses = await crud.get_user_data(db, user.id, year, month_num)

    transactions = []
    for inc in incomes:
        transactions.append({
            "date": inc.created_at,
            "description": inc.description,
            "type": "income",
            "type_label": "Доход",
            "amount": float(inc.amount),
            "currency": inc.currency,
        })
    for exp in expenses:
        transactions.append({
            "date": exp.created_at,
            "description": exp.description,
            "type": "expense",
            "type_label": "Расход",
            "amount": float(exp.amount),
            "currency": exp.currency,
        })

    def sort_key(t):
        dt = t["date"]
        if dt is None:
            return 0
        try:
            return dt.timestamp()
        except Exception:
            return 0

    transactions.sort(key=sort_key, reverse=True)

    totals: Dict[str, Dict[str, float]] = {}
    for t in transactions:
        cur = t["currency"]
        totals.setdefault(cur, {"income": 0.0, "expense": 0.0, "balance": 0.0})
        totals[cur]["income" if t["type"] == "income" else "expense"] += t["amount"]
    for cur in totals:
        totals[cur]["balance"] = totals[cur]["income"] - totals[cur]["expense"]

    current_year = datetime.now().year
    years = list(range(current_year, current_year + 3))

    months_ru_tuple = [
        ("01", "Январь"), ("02", "Февраль"), ("03", "Март"),
        ("04", "Апрель"), ("05", "Май"), ("06", "Июнь"),
        ("07", "Июль"), ("08", "Август"), ("09", "Сентябрь"),
        ("10", "Октябрь"), ("11", "Ноябрь"), ("12", "Декабрь"),
    ]

    response = templates.TemplateResponse(
        request,
        "data.html",
        {
            "user": user,
            "transactions": transactions,
            "totals": totals,
            "year": year,
            "month_num": month_num,
            "date": date,
            "export_period": export_period,
            "years": years,
            "months_ru": months_ru_tuple,
        },
    )
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.get("/data/export")
async def export_data(
    request: Request,
    date: str = "",
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login")

    year  = date[:4]  if len(date) >= 4 else ""
    month = date[5:7] if len(date) >= 7 else ""
    incomes, expenses = await crud.get_user_data(db, user.id, year, month)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Транзакции"
    ws.append(["Дата", "Описание", "Тип", "Сумма", "Валюта"])

    rows = []
    for inc in incomes:
        rows.append((inc.created_at, inc.description, "Доход", float(inc.amount), inc.currency))
    for exp in expenses:
        rows.append((exp.created_at, exp.description, "Расход", float(exp.amount), exp.currency))

    def exp_key(r):
        dt = r[0]
        if dt is None:
            return 0
        try:
            return dt.timestamp()
        except Exception:
            return 0

    rows.sort(key=exp_key, reverse=True)

    for row in rows:
        ws.append([
            row[0].strftime("%d.%m.%Y %H:%M") if row[0] else "",
            row[1],
            row[2],
            row[3],
            row[4],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    if year and month:
        period = f"{year}-{month}"
    elif year:
        period = year
    else:
        period = "всё_время"

    filename = f"Финансовый_отчёт_{period}.xlsx"
    encoded = quote(filename, safe="")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


@app.get("/profile")
async def profile(request: Request, db: AsyncSession = Depends(get_db)):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login")
    _now = datetime.now()
    incomes, expenses = await crud.get_user_data(
        db, user.id, str(_now.year), f"{_now.month:02d}"
    )

    totals: Dict[str, Dict[str, float]] = {}
    for income in incomes:
        cur = income.currency
        totals.setdefault(cur, {"income": 0.0, "expense": 0.0, "balance": 0.0})
        totals[cur]["income"] += float(income.amount)
    for expense in expenses:
        cur = expense.currency
        totals.setdefault(cur, {"income": 0.0, "expense": 0.0, "balance": 0.0})
        totals[cur]["expense"] += float(expense.amount)
    for cur in totals:
        totals[cur]["balance"] = totals[cur]["income"] - totals[cur]["expense"]

    total_expenses_rub = totals.get("RUB", {}).get("expense", 0.0)
    limit = float(user.monthly_limit) if user.monthly_limit else 0.0
    spent_percent = min(round(total_expenses_rub / limit * 100) if limit > 0 else 0, 100)

    months_ru_dict = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }

    response = templates.TemplateResponse(
        request,
        "profile.html",
        {
            "user": user,
            "total_expenses_rub": total_expenses_rub,
            "spent_percent": spent_percent,
            "success": request.query_params.get("success"),
            "totals": totals,
            "month_name": months_ru_dict[datetime.now().month],
            "income_count": len(incomes),
            "expense_count": len(expenses),
        },
    )
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.post("/profile/limit")
async def update_limit(
    request: Request,
    monthly_limit: float = Form(...),
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)
    await crud.update_monthly_limit(db, user.id, monthly_limit)
    response = RedirectResponse("/profile?success=1", status_code=302)
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response


@app.get("/analytics")
async def analytics(
    request: Request,
    date: str = "",
    currency: str = "RUB",
    db: AsyncSession = Depends(get_db),
):
    user, new_tokens = await get_user(request, db)
    if not user:
        return RedirectResponse("/login")

    year  = date[:4]  if len(date) >= 4 else ""
    month = date[5:7] if len(date) >= 7 else ""
    incomes, expenses = await crud.get_user_data(db, user.id, year, month)

    incomes_converted = special_funcs.convert_all_to_currency([i.__dict__ for i in incomes], currency)
    expenses_converted = special_funcs.convert_all_to_currency([e.__dict__ for e in expenses], currency)

    income_categories = special_funcs.categorize_transactions(incomes_converted)
    expense_categories = special_funcs.categorize_transactions(expenses_converted)

    response = templates.TemplateResponse(
        request,
        "analytics.html",
        {
            "user": user,
            "income_categories": income_categories,
            "expense_categories": expense_categories,
            "currency": currency,
        },
    )
    if new_tokens:
        set_auth_cookies(response, new_tokens)
    return response
